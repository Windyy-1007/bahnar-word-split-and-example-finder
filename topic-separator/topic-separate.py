import pandas as pd
import xlsxwriter as xlsx
import openpyxl as pxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

nameFile = 'Kinh Thanh Ba Na_578-660.xlsx'
inputfile = r'KinhThanh/' + nameFile
sepTopicType = 5
outputfileXLSX = r'MERGE/' + r'Sepearated_' + nameFile 

# Type -1: Hopeless: Input: Bahnar, Vietnamese --> Output: Bahnar, Vietnamese, linkSource, '0' 
# Type 0: Input: Chapters, Num, Bahnar, Vietnamese --> Output: Bahnar, Vietnamese, linkSource, Chapter
# Type 1: Input: Bahnar, Vietnamese (chapters are in bold) --> Output: Bahnar, Vietnamese, linkSource, Chapter
# Type 2: Input: Bahnar, Vietnamese (first word of Vietnamese is number) --> Output: Bahnar, Vietnamese, linkSource, Chapter
# Type 3: Input: Bahnar, Vietnamese (if cell1 is number and cell2 is blank, chapter is cell1) --> Output: Bahnar, Vietnamese, linkSource, Chapter
# Type 4: Input: Bahnar, Vietnamese (if cell1 is number and cell2 is number, chapter is cell1) --> Output: Bahnar, Vietnamese, linkSource, Chapter
# Type 5: Imput: Num, Bahnar, Vietnamese --> Output: Bahnar, Vietnamese, linkSource, Chapter

def check_space(word):
    if word[0] == ' ':
        return word[1:]
    return word

def remove_first_word(word):
    return word.split(' ', 1)[1]

def dataFromXLSX_typeHopeless(file):
    data = [
        ['A', 'B', 'C', 'D'],
    ]
    df1 = pd.read_excel(file)
    for row in range(df1.shape[0]):
        cell1 = str(df1.iat[row, 0])
        cell2 = str(df1.iat[row, 1])
        if not cell1.strip() or cell1 == 'nan':
            cell1 = ' '
        if not cell2.strip() or cell2 == 'nan':
            cell2 = ' '
        data.append(['0', '-', cell1, cell2])
    return data

def dataFromXLSX_type0(file):
    # Data has 4 columns: Chapter, Num, Bahnar, Vietnamese.
    data = [
        ['A', 'B', 'C', 'D'],
    ]
    df1 = pd.read_excel(file)
    # For each row, check 4 cells
    for row in range(df1.shape[0]):
        for col in range(df1.shape[1] - 3):
            if df1.iat[row, col+1] == None:
                continue
            # Make sure cell 1, 2, 3, 4 are strings
            cell1 = str(df1.iat[row, col])
            cell2 = str(df1.iat[row, col+1])
            cell3 = str(df1.iat[row, col+2])
            cell4 = str(df1.iat[row, col+3])
            if not cell1.strip() or cell1 == 'nan':
                if row == 0:
                    cell1 = 'Chapter 1'
                else:
                    cell1 = '-'
            # If cell2, 3, 4 is empty, add "-" to cell2, 3, 4
            if not cell2.strip() or cell2 == 'nan':
                cell2 = ' '
            if not cell3.strip() or cell3 == 'nan':
                cell3 = ' '
            if not cell4.strip() or cell4 == 'nan':
                cell4 = ' '
            # Add row's content to data
            cell1 = check_space(cell1)
            cell2 = check_space(cell2)
            cell3 = check_space(cell3)
            cell4 = check_space(cell4)
            data.append([cell1, cell2, cell3, cell4])
    return data

def dataFromXLSX_type1(file):
    data=[
        ['A', 'B', 'C', 'D'],
    ]
    #file only have 2 columns: Bahnar, Vietnamese (cell1, cell2)
    # Chapters are in bold. If cell1 is bold, chapter = cell1, else chapter = previous chapter
    df1 = pxl.load_workbook(file)
    sheet = df1.active
    for row in sheet.iter_rows(values_only=False):
        cell1 = str(row[0].value)
        cell2 = str(row[1].value)
        if not cell1.strip() or cell1 == 'nan':
            cell1 = ' '
        if not cell2.strip() or cell2 == 'nan':
            cell2 = ' '
        # If cell1 is bold, cell1 is chapter, else cell1 is empty
        if row[0].font.bold:
            chapter = cell1
        else:
            chapter = '-'
        
        data.append([chapter, '-', cell1, cell2])
    return data

def dataFromXLSX_type2(file):
    data=[
        ['A', 'B', 'C', 'D'],
    ]
    #file only have 2 columns: Bahnar, Vietnamese (cell1, cell2)
    # First word of Vietnamese is number. If cell2 is number, chapter = cell2, else chapter = previous chapter
    df1 = pxl.load_workbook(file)
    sheet = df1.active
    for row in sheet.iter_rows(values_only=False):
        cell1 = str(row[0].value)
        cell2 = str(row[1].value)
        if not cell1.strip() or cell1 == 'nan':
            cell1 = ' '
        if not cell2.strip() or cell2 == 'nan':
            cell2 = ' '
        # If first word of cell2 is '1' --> chapter = num_row, else chapter = previous chapter
        if cell2[0] == '1' and not cell2[1].isdigit():
            # Chapter is the current number of row
            chapter = str(row[0].row)
        else:
            chapter = '-'
            
        # If first word of cell1 or cell2 is number, remove it
        if cell1[0].isdigit():
            cell1 = remove_first_word(cell1)
        if cell2[0].isdigit():
            cell2 = remove_first_word(cell2)
        
        data.append([chapter, '-', cell1, cell2])
    return data

def dataFromXLSX_type3(file):
    data = [
        ['A', 'B', 'C', 'D'],
    ]
    
    df1 = pd.read_excel(file)
    for row in range(df1.shape[0]):
        cell1 = str(df1.iat[row, 0])
        cell2 = str(df1.iat[row, 1])
        if not cell1.strip() or cell1 == 'nan':
            cell1 = ' '
        if not cell2.strip() or cell2 == 'nan':
            cell2 = ' '
        if (cell1[0] in '0123456789') and (not cell2.strip() or cell2 == 'nan'):
            chapter = cell1
        else:
            chapter = '-'
        data.append([chapter, '-', cell1, cell2])
        print(data[row])
    return data

def dataFromXLSX_type4(file):
    data = [
        ['A', 'B', 'C', 'D'],
    ]
    
    df1 = pd.read_excel(file)
    for row in range(df1.shape[0]):
        cell1 = str(df1.iat[row, 0])
        cell2 = str(df1.iat[row, 1])
        if not cell1.strip() or cell1 == 'nan':
            cell1 = ' '
        if not cell2.strip() or cell2 == 'nan':
            cell2 = ' '
        # If cell1 and cell2 only contain numbers, chapter = cell1
        if cell1.isdigit() and cell2.isdigit():
            chapter = cell1
        else:
            chapter = '-'
        data.append([chapter, '-', cell1, cell2])
        print(data[row])
    return data

def dataFromXLSX_type5(file):
    data=[
        ['A', 'B', 'C', 'D'],
    ]
    
    #If cell1.value = 1, chapter is cell2.value, else chapter is previous chapter
    df1 = pxl.load_workbook(file)
    sheet = df1.active
    for row in sheet.iter_rows(values_only=False):
        cell1 = str(row[0].value)
        cell2 = str(row[1].value)
        cell3 = str(row[2].value)
        if not cell1.strip() or cell1 == 'None':
            cell1 = ' '
        if not cell2.strip() or cell2 == 'None' or not cell3.strip() or cell3 == 'None':
            continue
        if cell1 == '1':
            chapter = cell2
        else:
            chapter = '-'
        data.append([chapter, '-', cell2, cell3])
    return data

    
            
def processData(data = None):
    if data == None:
        return

    for row in range(1, len(data)):
        if data[row][0] == '-':
            data[row][0] = data[row-1][0]
    
    result = [
        ['A', 'B', 'C', 'D'],
    ]
    # Data also has cell1, cell2, cell3, cell4
    # Append in this order: cell3, cell4, linkSource, cell1
    linkSource = r'Gia Lai\Kinh Thánh Ba Na - Số hóa\\' + nameFile
    for row in range(1, len(data)):
        if((data[row][2] == '' and data[row][3] == '') or data[row][2] == 'None' or data[row][3] == 'None'):
            continue
        result.append([data[row][2], data[row][3], linkSource, data[row][0]])
    return result

def dataToXLSX(data, outputfile):
    # If file not exist, create new file
    if not os.path.exists(outputfile):
        workbook = xlsx.Workbook(outputfile)
        workbook
        workbook.close()
    
    workbook = xlsx.Workbook(outputfile)
    worksheet = workbook.add_worksheet()
    for row in range(len(data)):
        for col in range(len(data[row])):
            worksheet.write(row, col, data[row][col])
    workbook.close()
    return

def main():
    if(sepTopicType == 0):
        data = dataFromXLSX_type0(inputfile)
    elif(sepTopicType == 1):
        data = dataFromXLSX_type1(inputfile)
    elif(sepTopicType == 2):
        data = dataFromXLSX_type2(inputfile)
    elif(sepTopicType == 3):
        data = dataFromXLSX_type3(inputfile)
    elif(sepTopicType == 4):
        data = dataFromXLSX_type4(inputfile)
    elif(sepTopicType == 5):
        data = dataFromXLSX_type5(inputfile)
    else:
        data = dataFromXLSX_typeHopeless(inputfile)
    result = processData(data)
    dataToXLSX(result, outputfileXLSX)
    
main()
    